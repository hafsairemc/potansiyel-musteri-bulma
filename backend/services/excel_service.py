import os
import tempfile
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.config import settings
from services.search_result_filter_service import result_conditions


class ExcelDisaAktarimServisi:
    sutunlar = [
        "Firma", "Website", "Ülke", "Şehir", "Email", "E-posta Durumu", "E-posta Kaynağı", "Telefon",
        "Kategori", "Eşleşen Terimler", "Sektör Eşleşmesi", "Müşteri Tipi",
        "Genel Skor", "İlgililik Skoru", "Alıcı Skoru", "Güven Skoru",
        "Gerekçe", "Kaynak Türü", "Platform", "Arama Kelimesi",
    ]
    columns = sutunlar

    def excele_aktar(self, data: list[dict], file_path: str) -> str:
        guvenli_veri = [
            {
                anahtar: f"'{deger}" if isinstance(deger, str) and deger.startswith(("=", "+", "-", "@")) else deger
                for anahtar, deger in satir.items()
            }
            for satir in data
        ]
        df = pd.DataFrame(guvenli_veri, columns=self.sutunlar)
        df.to_excel(file_path, index=False, engine="openpyxl")

        calisma_kitabi = openpyxl.load_workbook(file_path)
        sayfa = calisma_kitabi.active
        sayfa.title = "Firma Listesi"

        dolgu = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        kenarlik = Border(*(Side(style="thin", color="D9D9D9") for _ in range(4)))

        for hucre in sayfa[1]:
            hucre.fill = dolgu
            hucre.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            hucre.border = kenarlik

        for satir in sayfa.iter_rows(min_row=2):
            for hucre in satir:
                hucre.border = kenarlik
                hucre.alignment = Alignment(vertical="center", wrap_text=True)

        for sutun_hucreleri in sayfa.columns:
            uzunluklar = [len(str(hucre.value)) for hucre in sutun_hucreleri if hucre.value is not None]
            sayfa.column_dimensions[get_column_letter(sutun_hucreleri[0].column)].width = min(max(uzunluklar or [10]) + 3, 50)

        sayfa.freeze_panes = "A2"
        sayfa.auto_filter.ref = sayfa.dimensions
        calisma_kitabi.save(file_path)
        return file_path

    export_to_excel = excele_aktar


ExcelExportService = ExcelDisaAktarimServisi


def toplu_aktarim_olustur(export_id: str) -> None:
    from core.database import SessionLocal
    from models.crawler_model import CrawlerCompany, CrawlerSearchJob, CrawlerSearchResult, SearchExport
    from services.db import admin_supabase

    db = SessionLocal()
    aktarim = db.query(SearchExport).filter(SearchExport.id == export_id).first()
    if not aktarim:
        db.close()
        return

    try:
        aktarim.status = "RUNNING"
        db.commit()

        satirlar = db.query(CrawlerSearchResult, CrawlerCompany, CrawlerSearchJob).join(
            CrawlerCompany, CrawlerCompany.id == CrawlerSearchResult.company_id
        ).join(CrawlerSearchJob, CrawlerSearchJob.id == CrawlerSearchResult.search_job_id).filter(
            *result_conditions(aktarim.batch_id, aktarim.filters or {}),
            CrawlerSearchJob.user_id == aktarim.user_id,
        ).order_by(CrawlerSearchResult.score.desc()).all()

        veri = [{
            "Firma": sirket.name or "Bilinmiyor", "Website": sonuc.source_url,
            "Ülke": sirket.country or islem.target_country or "", "Şehir": sirket.city or "",
            "Email": sirket.email or "",
            "E-posta Durumu": "Kamusal kaynakta bulundu" if sirket.email_status == "public_source" else ("Doğrulandı" if sirket.email_status == "verified" else ""),
            "E-posta Kaynağı": sirket.email_source_url or "", "Telefon": sirket.phone or "",
            "Kategori": sonuc.category_path or "", "Eşleşen Terimler": ", ".join(sonuc.matched_terms or []),
            "Sektör Eşleşmesi": sonuc.sector_match, "Müşteri Tipi": sonuc.customer_type or "",
            "Genel Skor": sonuc.score, "İlgililik Skoru": sonuc.relevance_score,
            "Alıcı Skoru": sonuc.buyer_score, "Güven Skoru": sonuc.confidence_score,
            "Gerekçe": sonuc.match_reason or "", "Kaynak Türü": sonuc.source,
            "Platform": sonuc.platform or "",
            "Arama Kelimesi": sonuc.search_query or islem.search_query,
        } for sonuc, sirket, islem in satirlar]

        yerel_yol = os.path.join(tempfile.gettempdir(), f"pusula-{aktarim.id}.xlsx")
        ExcelDisaAktarimServisi().excele_aktar(veri, yerel_yol)

        dosya_yuklendi = False
        if admin_supabase and getattr(settings, "reports_bucket", None):
            try:
                nesne_adi = f"{aktarim.user_id}/{aktarim.id}.xlsx"
                with open(yerel_yol, "rb") as dosya_akisi:
                    admin_supabase.storage.from_(settings.reports_bucket).upload(
                        nesne_adi, dosya_akisi, {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "upsert": "true"}
                    )
                aktarim.file_url = f"storage:{nesne_adi}"
                dosya_yuklendi = True
                try:
                    os.remove(yerel_yol)
                except Exception:
                    pass
            except Exception:
                dosya_yuklendi = False

        if not dosya_yuklendi:
            aktarim.file_url = yerel_yol

        aktarim.status = "COMPLETED"
        db.commit()
    except Exception as exc:
        db.rollback()
        aktarim = db.query(SearchExport).filter(SearchExport.id == export_id).first()
        if aktarim:
            aktarim.status = "FAILED"
            aktarim.error_message = str(exc)[:2000]
            db.commit()
    finally:
        db.close()


create_batch_export = toplu_aktarim_olustur
